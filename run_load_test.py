#!/usr/bin/env python3
import asyncio
import time
import os
import sys
import statistics
from datetime import datetime

# Import external packages. We will verify they are installed in the script.
try:
    import aiohttp
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
except ImportError:
    print("❌ ERROR: Missing dependencies. Please run:")
    print("   pip3 install aiohttp openpyxl")
    sys.exit(1)

# Configuration
TARGET_URL = "https://web-production-86613.up.railway.app/"
CONCURRENT_USERS = 100
DURATION_SECONDS = 60

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_EXCEL = os.path.join(SCRIPT_DIR, "reports", "Load_Test_Report.xlsx")

# Statistics trackers
latencies = []
status_codes = {}
total_requests = 0
successful_requests = 0
failed_requests = 0

async def send_request(session, worker_id):
    global total_requests, successful_requests, failed_requests
    start_time = time.perf_counter()
    try:
        async with session.get(TARGET_URL) as response:
            latency = (time.perf_counter() - start_time) * 1000  # Convert to ms
            status = response.status
            latencies.append(latency)
            status_codes[status] = status_codes.get(status, 0) + 1
            total_requests += 1
            if 200 <= status < 300:
                successful_requests += 1
            else:
                failed_requests += 1
    except Exception as e:
        latency = (time.perf_counter() - start_time) * 1000
        latencies.append(latency)
        status_codes["Error"] = status_codes.get("Error", 0) + 1
        total_requests += 1
        failed_requests += 1

async def worker(session, worker_id, stop_event):
    while not stop_event.is_set():
        await send_request(session, worker_id)
        # Avoid tight loop overloading local CPU by yielding control
        await asyncio.sleep(0.01)

async def run_load_test():
    print(f"🚀 Starting Load Test against: {TARGET_URL}")
    print(f"👥 Concurrent Users: {CONCURRENT_USERS}")
    print(f"⏱️  Duration: {DURATION_SECONDS} seconds")
    print("--------------------------------------------------")

    stop_event = asyncio.Event()
    
    # Track RPS over time (seconds)
    rps_timeline = []

    async with aiohttp.ClientSession() as session:
        # Spawn the workers
        tasks = [asyncio.create_task(worker(session, i, stop_event)) for i in range(CONCURRENT_USERS)]
        
        start_time = time.time()
        last_check_time = start_time
        last_request_count = 0

        # Run test for duration
        while time.time() - start_time < DURATION_SECONDS:
            await asyncio.sleep(1)
            current_time = time.time()
            elapsed = current_time - last_check_time
            current_total = total_requests
            reqs_in_interval = current_total - last_request_count
            interval_rps = reqs_in_interval / elapsed if elapsed > 0 else 0
            
            rps_timeline.append((int(current_time - start_time), round(interval_rps, 1)))
            print(f"⏱️  {int(current_time - start_time)}s elapsed | Total Requests: {current_total} | Current RPS: {interval_rps:.1f} req/sec")
            
            last_check_time = current_time
            last_request_count = current_total

        # Stop workers
        stop_event.set()
        await asyncio.gather(*tasks, return_exceptions=True)

    print("--------------------------------------------------")
    print("✅ Load Test Completed!")
    
    # Calculate stats
    total_duration = time.time() - start_time
    avg_rps = total_requests / total_duration if total_duration > 0 else 0
    
    if latencies:
        avg_latency = statistics.mean(latencies)
        min_latency = min(latencies)
        max_latency = max(latencies)
        latencies.sort()
        p50 = statistics.median(latencies)
        p95 = latencies[int(len(latencies) * 0.95)] if len(latencies) >= 20 else latencies[-1]
        p99 = latencies[int(len(latencies) * 0.99)] if len(latencies) >= 100 else latencies[-1]
    else:
        avg_latency = min_latency = max_latency = p50 = p95 = p99 = 0

    print(f"📊 Summary:")
    print(f"   • Total Requests: {total_requests}")
    print(f"   • Success Rate  : {(successful_requests/total_requests*100):.1f}%" if total_requests else "N/A")
    print(f"   • Avg RPS       : {avg_rps:.1f} req/sec")
    print(f"   • Response Time : Avg: {avg_latency:.1f}ms | Min: {min_latency:.1f}ms | Max: {max_latency:.1f}ms")
    print(f"   • Percentiles   : p50: {p50:.1f}ms | p95: {p95:.1f}ms | p99: {p99:.1f}ms")

    # Post markdown summary to GitHub Actions if available
    summary_path = os.environ.get("GITHUB_STEP_SUMMARY")
    if summary_path:
        success_rate = (successful_requests / total_requests * 100) if total_requests else 0
        try:
            with open(summary_path, "w", encoding="utf-8") as sf:
                sf.write("# ⚡ Digipay Backend Load Test Report\n\n")
                sf.write(f"> **Target URL:** `{TARGET_URL}`\n")
                sf.write(f"> **Environment:** Railway Hosted Instance\n")
                sf.write(f"> **Run Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} UTC\n\n")
                sf.write("## 📊 Performance Summary\n\n")
                sf.write("| Metric | Value |\n|---|---|\n")
                sf.write(f"| 👥 Concurrent Users (VUs) | **{CONCURRENT_USERS}** |\n")
                sf.write(f"| ⏱️ Test Duration | **{total_duration:.1f}s** |\n")
                sf.write(f"| 🔢 Total Requests Sent | **{total_requests}** |\n")
                sf.write(f"| ✅ Success Rate | **{success_rate:.1f}%** |\n")
                sf.write(f"| ⚡ Average Throughput | **{avg_rps:.1f} req/sec** |\n")
                sf.write(f"| ⏱️ Average Latency | **{avg_latency:.1f} ms** |\n")
                sf.write(f"| ⏱️ Min Latency | **{min_latency:.1f} ms** |\n")
                sf.write(f"| ⏱️ Max Latency | **{max_latency:.1f} ms** |\n")
                sf.write(f"| 📊 p50 (Median) | **{p50:.1f} ms** |\n")
                sf.write(f"| 📊 p95 | **{p95:.1f} ms** |\n")
                sf.write(f"| 📊 p99 | **{p99:.1f} ms** |\n\n")
                sf.write("---\n_📥 Download the full styled `.xlsx` report from the **Artifacts** section._\n")
            print("📝 GITHUB_STEP_SUMMARY updated successfully.")
        except Exception as e:
            print(f"⚠️ Failed to write to GITHUB_STEP_SUMMARY: {e}")

    # Generate Excel Report
    generate_excel_report(total_duration, avg_rps, avg_latency, min_latency, max_latency, p50, p95, p99, rps_timeline)

def generate_excel_report(total_duration, avg_rps, avg_latency, min_latency, max_latency, p50, p95, p99, rps_timeline):
    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
    wb = openpyxl.Workbook()
    
    # Style Dashboard Sheet
    ws = wb.active
    ws.title = "📊 Load Test Results"
    ws.sheet_view.showGridLines = False
    
    # Colors
    DARK_BG = "1F2937"    # Dark grey
    ACCENT = "3B82F6"     # Blue accent
    ALT_ROW = "F3F4F6"
    BORDER_C = "D1D5DB"
    
    thin = Side(style="thin", color=BORDER_C)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Title Block
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "🚀 DIGIPAY BACKEND LOAD TEST REPORT"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=DARK_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Metadata Block
    ws.merge_cells("A2:E2")
    meta_cell = ws["A2"]
    meta_cell.value = f"Run Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Environment: Railway Hosted Backend"
    meta_cell.font = Font(name="Calibri", italic=True, size=10, color="FFFFFF")
    meta_cell.fill = PatternFill("solid", fgColor="4B5563")
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Setup Metrics Table Headers
    headers = ["Metric / Parameter", "Value", "", "RPS Timeline (Seconds)", "RPS Value"]
    ws.row_dimensions[4].height = 24
    for col_idx, h in enumerate(headers, start=1):
        if h == "":
            continue
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        
    metrics = [
        ("Target URL", TARGET_URL),
        ("Concurrent Users (VUs)", CONCURRENT_USERS),
        ("Test Duration (Seconds)", round(total_duration, 1)),
        ("Total Requests Sent", total_requests),
        ("Successful Requests", successful_requests),
        ("Failed Requests", failed_requests),
        ("Success Rate", f"{(successful_requests/total_requests*100):.1f}%" if total_requests else "0%"),
        ("Average Throughput (RPS)", round(avg_rps, 1)),
        ("Min Response Time (Latency)", f"{min_latency:.1f} ms"),
        ("Average Response Time", f"{avg_latency:.1f} ms"),
        ("Max Response Time", f"{max_latency:.1f} ms"),
        ("p50 (Median) Latency", f"{p50:.1f} ms"),
        ("p95 Latency", f"{p95:.1f} ms"),
        ("p99 Latency", f"{p99:.1f} ms")
    ]
    
    # Write Metrics
    for idx, (m, v) in enumerate(metrics, start=5):
        c1 = ws.cell(row=idx, column=1, value=m)
        c2 = ws.cell(row=idx, column=2, value=v)
        c1.border = cell_border
        c2.border = cell_border
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="right", vertical="center")
        c1.font = Font(name="Calibri", bold=True if "Latency" in m or "Rate" in m else False)
        c2.font = Font(name="Calibri", bold=True)
        if idx % 2 == 0:
            c1.fill = PatternFill("solid", fgColor=ALT_ROW)
            c2.fill = PatternFill("solid", fgColor=ALT_ROW)
            
    # Write Timeline
    for idx, (sec, val) in enumerate(rps_timeline, start=5):
        c1 = ws.cell(row=idx, column=4, value=sec)
        c2 = ws.cell(row=idx, column=5, value=val)
        c1.border = cell_border
        c2.border = cell_border
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c2.alignment = Alignment(horizontal="right", vertical="center")
        if idx % 2 == 0:
            c1.fill = PatternFill("solid", fgColor=ALT_ROW)
            c2.fill = PatternFill("solid", fgColor=ALT_ROW)
            
    # Set Widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 16
    
    # Save Report
    wb.save(OUTPUT_EXCEL)
    print(f"📊 Styled Excel report saved successfully to: {OUTPUT_EXCEL}")

if __name__ == "__main__":
    asyncio.run(run_load_test())
